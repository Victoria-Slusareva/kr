import sys
from ui_interface import *

from Custom_Widgets.Widgets import *
from connection import get_connection, get_data_list
from PySide2.QtWidgets import QTableWidgetItem,  QApplication, QPushButton, QToolTip, QFileDialog
from PySide2.QtCharts import QtCharts
from PySide2.QtGui import QPainter, QCursor, QColor
import numpy as np
from openpyxl import Workbook

settings = QSettings()

class MainWindow(QMainWindow):
    def __init__(self, parent=None):
        QMainWindow.__init__(self)
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)

        loadJsonStyle(self, self.ui, jsonFiles = {"style.json"}) 

        self.show() 

        QAppSettings.updateAppSettings(self)

        self.validate()

        self.ui.refreshBtn.clicked.connect(self.refresh)
        self.ui.searchBtn.clicked.connect(self.search_result)

        self.ui.searchTableNameBtn.clicked.connect(self.search_table)
        self.ui.clearTableSearchBtn.clicked.connect(self.clear_search_table)

        self.tables_group.buttonClicked[int].connect(self.on_button_clicked)

        self.ui.mainParameterSet.clicked.connect(self.set_main_param)
        self.ui.mainParameterValueSet.clicked.connect(self.set_main_param_val)
        self.ui.agregationTypeSet.clicked.connect(self.set_agregation_type)
        self.ui.agregationParameterSet.clicked.connect(self.set_agregation_parameter)
        self.ui.mainParameterReset.clicked.connect(self.reset_main_param)
        self.ui.mainParameterValueReset.clicked.connect(self.reset_main_param_val)
        self.ui.agregationTypeReset.clicked.connect(self.reset_agregation_type)
        self.ui.agregationParameterReset.clicked.connect(self.reset_agregation_parameter)
        self.ui.continueLineChartBtn.clicked.connect(self.draw_graph)
        self.ui.returnToChartParameters.clicked.connect(self.line_chart_params)
        self.ui.linePlotBtn.clicked.connect(self.line_chart_params)
        self.ui.refreshChartBtn.clicked.connect(self.change_plot_type)

        self.ui.ratingMainParameterSet.clicked.connect(self.set_rating_main_param)
        self.ui.ratingYearSet.clicked.connect(self.set_rating_year)
        self.ui.ratingAgregationTypeSet.clicked.connect(self.set_rating_agregation_type)
        self.ui.ratingAgregationParameterSet.clicked.connect(self.set_rating_agregation_parameter)
        self.ui.mainParameterReset.clicked.connect(self.reset_rating_main_param)
        self.ui.ratingYearReset.clicked.connect(self.reset_rating_year)
        self.ui.ratingAgregationTypeReset.clicked.connect(self.reset_rating_agregation_type)
        self.ui.ratingAgregationParameterReset.clicked.connect(self.reset_rating_agregation_parameter)
        self.ui.continueRatingBtn.clicked.connect(self.draw_rating)
        self.ui.returnToRatingParameters.clicked.connect(self.rating_chart_params)
        self.ui.saveTableBtn.clicked.connect(self.handleSave)
        self.ui.ratingSaveBtn.clicked.connect(self.save_rating_chart)
        self.ui.ratingPlotBtn.clicked.connect(self.rating_chart_params)
        self.ui.refreshRatingBtn.clicked.connect(self.redraw_rating)

    def validate(self):
        
        self.dbname = 'games'
        self.tablename = 'game'
        try:
            # trying connect to table
            self.ui.homeLbl.setText(self.tablename)
            self.setup_table(self.tablename, self.dbname)
            self.get_rating_year()
            self.show_ui(table=True)
            
        except Exception as e:
            print(e)

    def get_rating_year(self):
        query = 'with get_year as( \
                    select g.id as game_id, \
                    extract(year from g.release_date) as year \
                    from game g \
                    ) \
                    select distinct \
                    y.year \
                    from game g \
                    left join get_year y on y.game_id=g.id \
                    order by y.year'
        self.year_parameters = get_data_list(get_connection(self.dbname), query)
        self.year_parameters = list(self.year_parameters[0])
        for i in range(len(self.year_parameters)):
            self.ui.ratingYearFrom.addItem(str(self.year_parameters[i][0]))
            self.ui.ratingYearTo.addItem(str(self.year_parameters[i][0]))

    def handleSave(self):
        path = QFileDialog.getSaveFileName(
                self, 'Save File', '','Excel files (*.xlsx)')
        if path:
            print("saving", path[0])
            wb = Workbook()
            sheet = wb.active
            for column in range(self.ui.tableWidget.columnCount()):
                header = str(self.ui.tableWidget.horizontalHeaderItem(column).text())
                sheet.cell(1, column+1, header)
                for row in range(self.ui.tableWidget.rowCount()):
                    text = str(self.ui.tableWidget.item(row, column).text())
                    sheet.cell(row+2, column+1, text)
            wb.save(path[0])
            wb.close()
        
    def show_ui(self,table = False):
        
        if table:
            self.show_table()
            self.show_ui()
            self.search_form()
            self.get_table_btn()
      
            
    def setup_table(self, tablename, dbname):

        query = 'select * from ' + str(tablename)
        self.table = get_data_list(get_connection(dbname), query)
        self.cols = list(self.table[1])  
        self.records = list(self.table[0])  
        
        
    def show_table(self):
        
        self.ui.tableWidget.setColumnCount(len(self.cols)) 
        self.ui.tableWidget.setHorizontalHeaderLabels(self.cols)
        self.ui.tableWidget.setRowCount(len(self.records)) 
        row_count = 0
        for row in self.records:
            for i in range(len(self.cols)):
                self.ui.tableWidget.setItem(row_count, i,  QTableWidgetItem(str(row[i])))
            row_count += 1

    def refresh(self):
        self.setup_table(self.tablename, self.dbname)
        self.ui.tableWidget.clear()
        self.show_table()

    def clear_search_form(self):
        self.setup_table(self.tablename, self.dbname)
        self.ui.verticalLayout_8.removeItem(self.search_spacer)
        for i in reversed(range(self.ui.verticalLayout_8.count())): 
            self.ui.verticalLayout_8.itemAt(i).widget().setParent(None)

    def search_form(self):

        font = QFont()
        font.setPointSize(8)
        self.lines = []
        self.search_spacer = QSpacerItem(20, 40, QSizePolicy.Minimum, QSizePolicy.Expanding)

        for i in range(len(self.cols)):
            self.line = QLineEdit(self.ui.searchFrm)
            self.line.setObjectName('searchField'+str(i))
            self.line.setPlaceholderText(self.cols[i])
            self.line.setFont(font)
            self.ui.verticalLayout_8.addWidget(self.line)
            self.lines.append((self.line.placeholderText(), self.line))
        
        self.ui.verticalLayout_8.addItem(self.search_spacer)


    def search_result(self):
        query = 'select * from ' + str(self.tablename)
        not_nan = []

        for i in range(len(self.lines)):
            if self.lines[i][1].text():
                not_nan.append((self.lines[i][0], self.lines[i][1].text()))

        if len(not_nan) > 1:
            for i in range(len(not_nan)):
                if i == 0 and not_nan[i][0] == 'name':
                    query += ' where '
                    query += str(not_nan[i][0]) + " like '%" + str(not_nan[i][1]) + "%' and "
                elif i == 0:
                    query += ' where '
                    query += str(not_nan[i][0]) + ' = ' + str(not_nan[i][1]) + ' and '
                elif i == (len(not_nan)-1) and not_nan[i][0] == 'name':
                    query += str(not_nan[i][0]) + " like '%" + str(not_nan[i][1]) + "%'"
                elif i == (len(not_nan)-1):
                    query += str(not_nan[i][0]) + ' = ' + str(not_nan[i][1])
                elif not_nan[i][0] == 'name': 
                    query += str(not_nan[i][0]) + " like '%" + str(not_nan[i][1]) + "%' and "
                else: 
                    query += str(not_nan[i][0]) + ' = ' + str(not_nan[i][1]) + ' and '
        elif len(not_nan) == 1 and not_nan[0][0] == 'name':
            query += ' where ' + str(not_nan[0][0]) + " like '%" + str(not_nan[0][1]) + "%'"
        elif len(not_nan) == 1:
            query += ' where ' + str(not_nan[0][0]) + ' = ' + str(not_nan[0][1])
        else: 
            pass

        self.table = get_data_list(get_connection(self.dbname), query)
        self.records = list(self.table[0])
        self.ui.tableWidget.clear()
        self.show_table()

    
    def get_table_btn(self):
        query = 'show tables from ' + str(self.dbname)
        self.table_names = get_data_list(get_connection(self.dbname), query)
        self.table_names = list(self.table_names[0])
        icon1 = QIcon()
        icon1.addFile(u":/icons/Icons/layout.png", QSize(), QIcon.Normal, QIcon.Off)
        font2 = QFont()
        font2.setPointSize(10)
        font2.setBold(False)
        font2.setWeight(75)
        self.tables = []
        self.tables_group = QButtonGroup()
        self.table_spacer = QSpacerItem(20, 40, QSizePolicy.Minimum, QSizePolicy.Expanding)

        for i in range(len(self.table_names)):
            self.tablesBtn = QPushButton(self.ui.page)
            self.tablesBtn.setObjectName('tablesBtn'+str(i))
            self.tablesBtn.setFont(font2)
            self.tablesBtn.setCursor(QCursor(Qt.PointingHandCursor))
            self.tablesBtn.setIcon(icon1)
            self.tablesBtn.setText(self.table_names[i][0])
            self.tablesBtn.setIconSize(QSize(24, 24))
            self.ui.verticalLayout_12.addWidget(self.tablesBtn)
            self.tables_group.addButton(self.tablesBtn, i+1)
            self.tablesBtn.setStyleSheet('background-color: transparent; border: none; font-weight: normal;')
            self.tables.append(self.tablesBtn)
        self.ui.verticalLayout_13.addItem(self.table_spacer)
        self.ui.verticalLayout_12.addItem(self.table_spacer)
        self.ui.toolBox.setCurrentIndex(len(self.table_names)+2)


    
    def on_button_clicked(self, id):

        for button in self.tables_group.buttons():
            if button is self.tables_group.button(id):
                button.setStyleSheet('background-color: #1f232a;  \
                border-bottom: 3px solid #755D90;   \
                font-weight: bold;')
                self.change_table(button.text())
                self.ui.mainPages.setCurrentWidget(self.ui.homePage)
            else: 
                button.setStyleSheet('background-color: transparent; border: none; font-weight: normal;')


    def search_table(self):
        shw_btn = self.ui.search_table_line.text()
        print(shw_btn)
        
        for button in self.tables_group.buttons():
            print(button.text())
            if shw_btn in button.text():
                button.setVisible(True)
            else: 
                button.setVisible(False)

    def clear_search_table(self):

        self.ui.search_table_line.clear()
        for button in self.tables_group.buttons():
            button.setVisible(True)
         
            
    def get_tableWidget_data(self):
       
       recs, data = [], []
       for i in range(self.ui.tableWidget.rowCount()):
           for j in range(self.ui.tableWidget.columnCount()):
               recs.append(self.ui.tableWidget.item(i,j).text())
               
       n = self.ui.tableWidget.columnCount()      
       for i in range(0, len(recs),n):
           data.append(recs[i:i+n])
       return data


    def change_table(self, table):
        self.tablename = table
        self.clear_search_form()
        self.ui.homeLbl.setText(self.tablename)
        self.refresh()
        self.search_form()

    def set_main_param(self):
        parameter = str(self.ui.mainParameter.currentText())
        if parameter != '--основной параметр--':
            query = 'select distinct name from ' + parameter
            self.line_parameters = get_data_list(get_connection(self.dbname), query)
            self.line_parameters = list(self.line_parameters[0])
            for i in range(len(self.line_parameters)):
                self.ui.mainParameterValue.addItem(self.line_parameters[i][0])
            self.ui.mainParameterValue.setEnabled(True)

    def set_main_param_val(self):
        parameter = str(self.ui.mainParameterValue.currentText())
        if parameter != '--значение основного параметра--':
            self.ui.agregationType.setEnabled(True)

    def set_agregation_type(self):
        parameter = str(self.ui.agregationType.currentText())
        if parameter != '--способ расчета--':
            self.ui.agregationParameter.setEnabled(True)

    def set_agregation_parameter(self):
        parameter = str(self.ui.agregationParameter.currentText())
        if parameter != '--парамер для расчета--':
            self.ui.continueLineChartBtn.setEnabled(True)

    def reset_main_param(self):
        self.ui.mainParameter.setCurrentIndex(0)
        self.ui.mainParameterValue.setEnabled(False)
        self.ui.mainParameterValue.setCurrentIndex(0)
        self.ui.agregationType.setEnabled(False)
        self.ui.agregationType.setCurrentIndex(0)
        self.ui.agregationParameter.setEnabled(False)
        self.ui.agregationParameter.setCurrentIndex(0)
        self.ui.continueLineChartBtn.setEnabled(False)
        for i in reversed(range(len(self.line_parameters))):
            self.ui.mainParameterValue.removeItem(i+1)

    def reset_main_param_val(self):
        self.ui.mainParameterValue.setCurrentIndex(0)
        self.ui.agregationType.setEnabled(False)
        self.ui.agregationType.setCurrentIndex(0)
        self.ui.agregationParameter.setEnabled(False)
        self.ui.agregationParameter.setCurrentIndex(0)
        self.ui.continueLineChartBtn.setEnabled(False)

    def reset_agregation_type(self):
        self.ui.agregationType.setCurrentIndex(0)
        self.ui.agregationParameter.setEnabled(False)
        self.ui.agregationParameter.setCurrentIndex(0)
        self.ui.continueLineChartBtn.setEnabled(False)

    def reset_agregation_parameter(self):
        self.ui.agregationParameter.setCurrentIndex(0)
        self.ui.continueLineChartBtn.setEnabled(False)

    def draw_graph(self):
        parameter1 = str(self.ui.mainParameter.currentText())
        parameter2 = str(self.ui.mainParameterValue.currentText())
        parameter3 = str(self.ui.agregationType.currentText())
        parameter4 = str(self.ui.agregationParameter.currentText())
        parameter3_value = ''
        
        if parameter3 == 'Количество':
            parameter3_value = 'count'
        elif parameter3 == 'Среднее':
            parameter3_value = 'avg'
        elif parameter3 == 'Сумма':
            parameter3_value = 'sum'

        if parameter1 != 'game':
            query = 'with get_year as( \
                        select g.id as game_id, extract(year from g.release_date) as year \
                        from game g \
                    ) \
                    select y.year, \
                        '+parameter3_value+'('+parameter4+') as '\
                        +parameter3_value+'_'+parameter4+' \
                    from game g \
                    left join game_'+parameter1+' gd on g.id = gd.game_id \
                    left join '+parameter1+' d on gd.'+parameter1+'_id = d.id \
                    left join get_year y on y.game_id=g.id \
                    where d.name = "'+parameter2+'" \
                    group by y.year, d.id \
                    order by y.year'
        else:
            query = 'with get_year as( \
                    select g.id as game_id, extract(year from g.release_date) as year \
                    from game g \
                ) \
                select y.year, \
                    '+parameter3_value+'(g.name) as '\
                    +parameter3_value+'_game \
                from game g \
                left join get_year y on y.game_id=g.id \
                group by y.year \
                order by y.year'


        self.graph_data = get_data_list(get_connection(self.dbname), query)
        self.bar_plot()

    def change_plot_type(self):
        gr_type = str(self.ui.graphTypeQCB.currentText())
        self.ui.gridLayout_14.removeWidget(self.chartview)
        if gr_type == 'График':
            self.line_plot()
        elif gr_type == 'Гистограмма':
            self.bar_plot()


    def line_plot(self):
        columns = list(self.graph_data[1])  
        records = list(self.graph_data[0]) 

        self.series = QtCharts.QLineSeries()

        for row in records:
            print(row[0], row[1])
            self.series.append(int(row[0]), float(row[1]))

        self.series.setPointsVisible(True) 
        self.series.hovered.connect(self.onSeriesHoverd)
        pen = QPen()
        pen.setWidth(3)
        self.series.setPen(pen)
        self.series.setColor(QColor(173,133,219))
        self.show_line_chart()


    def show_line_chart(self):
        self.chart =  QtCharts.QChart()
 
        self.chart.legend().hide()
        self.chart.addSeries(self.series)
        self.chart.createDefaultAxes()

        self.chart.axisY().setMin(0)
        self.chart.axisX().setLabelFormat("%d")

        minx = int(self.names[0])
        maxx = int(self.names[-1])
        self.chart.axisX().setRange(minx, maxx)
 
        self.chartview = QtCharts.QChartView(self.chart)
        self.chartview.setRenderHint(QPainter.Antialiasing)
        self.chart.setAnimationOptions(QtCharts.QChart.SeriesAnimations)
        self.chartview.chart().setTheme(QtCharts.QChart.ChartThemeLight)

        sizePolicy = QSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        sizePolicy.setHeightForWidth(self.chartview.sizePolicy().hasHeightForWidth())
        self.chartview.setSizePolicy(sizePolicy)
        self.chartview.setMinimumSize(QSize(0,300))

        self.ui.gridLayout_14.addWidget(self.chartview)
        self.ui.chartWdg.setStyleSheet(u"background-color: transparent")
        
        

    def bar_plot(self):
        columns = list(self.graph_data[1])  
        records = list(self.graph_data[0]) 

        self.series = QtCharts.QBarSeries()
        set1 = QtCharts.QBarSet('Set1')
        set1.setColor(QColor(173,133,219))
        self.names = []

        for row in records:
            print(row[0], row[1])
            set1.append(float(row[1]))
            self.names.append(str(row[0]))
        self.series.append(set1)
        self.show_bar_chart()


    def show_bar_chart(self):
        self.chart =  QtCharts.QChart()
 
        self.chart.legend().hide()
        self.chart.addSeries(self.series)

        axis = QtCharts.QBarCategoryAxis()
        axis.append(self.names)
        self.chart.createDefaultAxes()
        self.chart.axisY().setMin(0)
        self.chart.setAxisX(axis, self.series)
 
 
        self.chartview = QtCharts.QChartView(self.chart)
        self.chartview.setRenderHint(QPainter.Antialiasing)
        self.chart.setAnimationOptions(QtCharts.QChart.SeriesAnimations)
        self.chartview.chart().setTheme(QtCharts.QChart.ChartThemeLight)

        sizePolicy = QSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        sizePolicy.setHeightForWidth(self.chartview.sizePolicy().hasHeightForWidth())
        self.chartview.setSizePolicy(sizePolicy)
        self.chartview.setMinimumSize(QSize(0,300))

        self.ui.gridLayout_14.addWidget(self.chartview)
        self.ui.chartWdg.setStyleSheet(u"background-color: transparent")
        
    def line_chart_params(self):
        self.ui.mainPages.setCurrentWidget(self.ui.chartParametersPage)
        self.ui.gridLayout_14.removeWidget(self.chartview)

    def onSeriesHoverd(self, point, state):
        if state:
            QToolTip.showText(QCursor.pos(), "x: %s\ny: %s" %
                        (int(point.x()), np.round(point.y(),2)))



    def set_rating_main_param(self):
        parameter = str(self.ui.ratingMainParameter.currentText())
        if parameter != '--основной параметр--':
            self.ui.ratingAgregationType.setEnabled(True)

    def set_rating_year(self):
        parameter1 = str(self.ui.ratingYearFrom.currentText())
        parameter2 = str(self.ui.ratingYearTo.currentText())

    def set_rating_agregation_type(self):
        parameter = str(self.ui.ratingAgregationType.currentText())
        if parameter != '--способ расчета--':
            self.ui.ratingAgregationParameter.setEnabled(True)

    def set_rating_agregation_parameter(self):
        parameter = str(self.ui.ratingAgregationParameter.currentText())
        if parameter != '--парамер для расчета--':
            self.ui.continueRatingBtn.setEnabled(True)

    def reset_rating_main_param(self):
        self.ui.ratingMainParameter.setCurrentIndex(0)
        self.ui.ratingAgregationType.setEnabled(False)
        self.ui.ratingAgregationType.setCurrentIndex(0)
        self.ui.ratingAgregationParameter.setEnabled(False)
        self.ui.ratingAgregationParameter.setCurrentIndex(0)
        self.ui.continueRatingBtn.setEnabled(False)

    def reset_rating_year(self):
        self.ui.ratingMainParameter.setCurrentIndex(0)

    def reset_rating_agregation_type(self):
        self.ui.ratingAgregationType.setCurrentIndex(0)
        self.ui.ratingAgregationParameter.setEnabled(False)
        self.ui.ratingAgregationParameter.setCurrentIndex(0)
        self.ui.continueRatingBtn.setEnabled(False)

    def reset_rating_agregation_parameter(self):
        self.ui.ratingAgregationParameter.setCurrentIndex(0)
        self.ui.continueRatingBtn.setEnabled(False)

    def draw_rating(self):
        parameter1 = str(self.ui.ratingMainParameter.currentText())
        parameter2_1 = str(self.ui.ratingYearFrom.currentText())
        parameter2_2 = str(self.ui.ratingYearTo.currentText())
        parameter3 = str(self.ui.ratingAgregationType.currentText())
        parameter4 = str(self.ui.ratingAgregationParameter.currentText())
        parameter5 = str(self.ui.ratingNumLE.text())
        print('P5\t\t',parameter5)
        
        if parameter5 == '':
            parameter5 = '10'

        parameter3_value = ''
        
        if parameter3 == 'Количество':
            parameter3_value = 'count'
        elif parameter3 == 'Среднее':
            parameter3_value = 'avg'
        elif parameter3 == 'Сумма':
            parameter3_value = 'sum'

        if parameter2_1 == '--с какого года--':
            parameter2_1 = str(self.year_parameters[0][0])
        if parameter2_2 == '--по какой год--':
            parameter2_2 = str(self.year_parameters[-1][0])

        
        if parameter1 != 'game':
            query = 'with get_year as( \
                        select g.id as game_id, extract(year from g.release_date) as year \
                        from game g \
                    ) \
                    select d.name as '+parameter1+'_name, \
                        '+parameter3_value+'('+parameter4+') as '\
                        +parameter3_value+'_'+parameter4+' \
                    from game g \
                    left join game_'+parameter1+' gd on g.id = gd.game_id \
                    left join '+parameter1+' d on gd.'+parameter1+'_id = d.id \
                    left join get_year y on y.game_id=g.id \
                    where y.year >= '+parameter2_1+' and y.year <= '+parameter2_2+' \
                    group by d.id \
                    order by '+parameter3_value+'_'+parameter4+' desc \
                    limit '+parameter5
        else:
            query = 'with get_year as( \
                        select g.id as game_id, extract(year from g.release_date) as year \
                        from game g \
                    ) \
                    select g.name as '+parameter1+'_name, \
                        '+parameter3_value+'('+parameter4+') as '\
                        +parameter3_value+'_'+parameter4+' \
                    from game g \
                    left join get_year y on y.game_id=g.id \
                    where y.year >= '+parameter2_1+' and y.year <= '+parameter2_2+' \
                    group by g.id \
                    order by '+parameter3_value+'_'+parameter4+' desc \
                    limit '+parameter5


        self.graph_data = get_data_list(get_connection(self.dbname), query)
        print(self.graph_data)
        self.rating_plot()


    def rating_plot(self):
        columns = list(self.graph_data[1])  
        records = list(self.graph_data[0]) 

        self.series = QtCharts.QBarSeries()
        set1 = QtCharts.QBarSet('Set1')
        set1.setColor(QColor(173,133,219))
        self.names = []

        for row in records:
            print(row[0], row[1])
            set1.append(float(row[1]))
            self.names.append(str(row[0]))
        self.series.append(set1)
        self.show_rating_chart()


    def show_rating_chart(self):
        self.chart =  QtCharts.QChart()
 
        self.chart.legend().hide()
        self.chart.addSeries(self.series)

        axis = QtCharts.QBarCategoryAxis()
        axis.append(self.names)
        self.chart.createDefaultAxes()
        self.chart.axisY().setMin(0)
        self.chart.setAxisX(axis, self.series)
 
 
        self.chartview1 = QtCharts.QChartView(self.chart)
        self.chartview1.setRenderHint(QPainter.Antialiasing)
        self.chart.setAnimationOptions(QtCharts.QChart.SeriesAnimations)
        self.chartview1.chart().setTheme(QtCharts.QChart.ChartThemeLight)

        sizePolicy = QSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        sizePolicy.setHeightForWidth(self.chartview1.sizePolicy().hasHeightForWidth())
        self.chartview1.setSizePolicy(sizePolicy)
        self.chartview1.setMinimumSize(QSize(0,300))

        self.ui.gridLayout_19.addWidget(self.chartview1)
        self.ui.chartWdg.setStyleSheet(u"background-color: transparent")

    def rating_chart_params(self):
        self.ui.mainPages.setCurrentWidget(self.ui.ratingParametersPage)
        self.ui.gridLayout_19.removeWidget(self.chartview1)

    def save_rating_chart(self):
        path = QFileDialog.getSaveFileName(
                self, 'Save File', '','Portable Network Graphics (*.png)')
        if path:
            print("saving", path[0])
            p = QPixmap(self.chartview1.grab())
            p.save(path[0])

    def redraw_rating(self):
        self.ui.gridLayout_19.removeWidget(self.chartview1)
        self.draw_rating()


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec_())
    
